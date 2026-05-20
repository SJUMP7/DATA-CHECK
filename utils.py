"""
utils.py — Hotel Contract AI Extractor
Powered by Anthropic Claude (claude-sonnet-4-6)
Replaces: google-genai / Gemini
"""

import base64
import json
import io
from datetime import datetime
import anthropic
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MODEL = "claude-sonnet-4-6"

# ── Column Schema ─────────────────────────────────────────────────────────────
# Columns A-F are system fields (filled by dashboard / user)
# Columns G-AQ are extracted from PDF
COLUMNS = [
    "_id", "hotel_id", "room_id", "status", "created_date", "edited_date",
    "start_date", "end_date", "refundable", "abf", "contract_type",
    "cutoff_date", "hotel_supplier", "important_message", "min_nights_stay",
    "min_advance_days", "net_price", "net_price_emerald", "net_price_ruby",
    "net_price_topaz", "promo_book_till", "promo_code", "promo_note",
    "room_allotment", "all_inclusive", "baby_cot", "cancellation_policy",
    "cancellation_policy_net", "early_check_in", "child_policy",
    "child_share_bed_abf", "child_extra_bed_abf", "extra_bed_abf",
    "extra_bed_no_abf", "full_board", "half_board", "hotel_extra_fees",
    "room_name", "hotel_transfer", "late_check_out", "meals_and_info",
    "tags", "action",
]

SYSTEM_COLS = COLUMNS[:6]   # A-F: do not extract
EXTRACT_COLS = COLUMNS[6:]  # G-AQ: AI fills these


# ── Client Helper ─────────────────────────────────────────────────────────────
def _client(api_key: str) -> anthropic.Anthropic:
    return anthropic.Anthropic(api_key=api_key)


def _pdf_block(pdf_bytes: bytes) -> dict:
    return {
        "type": "document",
        "source": {
            "type": "base64",
            "media_type": "application/pdf",
            "data": base64.standard_b64encode(pdf_bytes).decode(),
        },
    }


def _excel_to_text(excel_bytes: bytes) -> str:
    """Convert Excel to a readable text table for Claude context."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"=== Sheet: {sheet_name} ===")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    header = [str(c) if c is not None else "" for c in row]
                    lines.append("COLUMNS: " + " | ".join(header))
                else:
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in cells):
                        lines.append(f"ROW {i}: " + " | ".join(cells))
        return "\n".join(lines)
    except Exception as e:
        return f"[Could not parse Excel: {e}]"


# ── API Key Validation ─────────────────────────────────────────────────────────
def validate_api_key(api_key: str) -> tuple[bool, str]:
    """Returns (is_valid, message)."""
    if not api_key or not api_key.strip().startswith("sk-ant-"):
        return False, "API Key ไม่ถูกต้อง — ต้องขึ้นต้นด้วย sk-ant-"
    try:
        client = _client(api_key)
        client.messages.create(
            model=MODEL,
            max_tokens=10,
            messages=[{"role": "user", "content": "hi"}],
        )
        return True, "API Key ถูกต้อง ✓"
    except anthropic.AuthenticationError:
        return False, "API Key ไม่ถูกต้องหรือหมดอายุ"
    except anthropic.RateLimitError:
        return False, "โควต้า API เต็ม — กรุณาลองใหม่ภายหลัง"
    except Exception as e:
        return False, f"ไม่สามารถเชื่อมต่อ: {e}"


# ── Feature 1 — Audit / Recheck ───────────────────────────────────────────────
_AUDIT_SYSTEM = """You are an expert hotel contract auditor for a Thai travel agency.
You receive a hotel rate contract PDF and an Excel rate sheet for the same property.
Your task is to cross-check every data point and produce a structured Thai-language audit report.

Output format rules:
- Use markdown with clear section headers
- Rate each field as: 🟢 VERIFIED | 🟡 REVIEW | 🔴 FAIL
- For FAIL items, specify exact PDF value vs Excel value and corrective action
- End with: คะแนนความถูกต้อง X.X% and บทสรุป: [one line]
- Be precise and concise; avoid filler sentences"""

_FOCUS_PROMPTS = {
    "Net Price & Extra Beds": "Focus ONLY on: net_price, extra_bed_abf, extra_bed_no_abf, child_extra_bed_abf columns.",
    "Cancellation Policy": "Focus ONLY on: cancellation_policy, refundable columns.",
    "Child Policy": "Focus ONLY on: child_policy, child_share_bed_abf, child_extra_bed_abf, baby_cot columns.",
    "Period & Seasons": "Focus ONLY on: start_date, end_date, contract_type, cutoff_date, min_nights_stay columns.",
    "Meals & Info": "Focus ONLY on: abf, full_board, half_board, meals_and_info, hotel_extra_fees columns.",
    "All-in-One Full Scan": "Audit ALL columns from G to AQ comprehensively.",
}


def stream_recheck_analysis(
    pdf_bytes: bytes,
    excel_bytes: bytes,
    api_key: str,
    focus_list: list[str],
):
    """
    Generator that streams markdown audit text chunk by chunk.
    Mirrors the interface of the original Gemini stream_recheck_analysis.
    """
    focus_instructions = "\n".join(
        _FOCUS_PROMPTS.get(f, "") for f in focus_list if f in _FOCUS_PROMPTS
    )

    excel_text = _excel_to_text(excel_bytes)

    user_prompt = f"""Please audit this hotel contract PDF against the Excel rate sheet below.

{focus_instructions}

## Excel Rate Sheet Content
```
{excel_text}
```

Produce your audit report now in Thai, strictly following the output format in your system instructions."""

    client = _client(api_key)
    try:
        with client.messages.stream(
            model=MODEL,
            max_tokens=4096,
            system=_AUDIT_SYSTEM,
            messages=[
                {
                    "role": "user",
                    "content": [
                        _pdf_block(pdf_bytes),
                        {"type": "text", "text": user_prompt},
                    ],
                }
            ],
        ) as stream:
            for text in stream.text_stream:
                yield text
    except anthropic.AuthenticationError:
        yield "**API Key ไม่ถูกต้อง** — กรุณาตรวจสอบ API Key ใน Sidebar"
    except anthropic.RateLimitError:
        yield "**โควต้า API เต็ม** — กรุณาลองใหม่ภายหลัง"
    except Exception as e:
        yield f"**ไม่สามารถเชื่อมต่อ**: {e}"


# ── Feature 2 — Generate Excel from PDF ───────────────────────────────────────
_EXTRACT_SYSTEM = """You are a hotel contract data extraction specialist for a Thai travel agency.
You receive a hotel rate contract PDF and must extract structured data into JSON format.

CRITICAL RULES:
1. Return ONLY a valid JSON array — no markdown, no explanation, no code fences.
2. Each element = one row in the Excel sheet (one room type × one season period).
3. Use null for fields not mentioned in the contract.
4. Dates must be ISO 8601 strings: "YYYY-MM-DD"
5. Prices must be numbers (integers), not strings.
6. HTML fields (cancellation_policy, child_policy, meals_and_info, hotel_transfer) must use proper HTML tags: <p>, <strong>, <span style="color:#f44336;">, etc.
7. contract_type values: "Main Contract" | "Promotion" | "Group Rate"
8. refundable: false for non-refundable, true for refundable
9. abf: "Included" if breakfast included, "Excluded" if not, null if unclear
10. room_allotment: "Free Sales" | "On Request" | null
11. action: always "insert"
12. tags: always []"""

_EXTRACT_COLS_SCHEMA = {
    "start_date": "Season start date (YYYY-MM-DD)",
    "end_date": "Season end date (YYYY-MM-DD)",
    "refundable": "bool — false if non-refundable",
    "abf": "\"Included\" | \"Excluded\" — breakfast status",
    "contract_type": "\"Main Contract\" | \"Promotion\" | \"Group Rate\"",
    "cutoff_date": "string — number of days e.g. \"7\"",
    "hotel_supplier": "Hotel name as string",
    "important_message": "string | null — important note from contract",
    "min_nights_stay": "string | null — e.g. \"3\"",
    "min_advance_days": "string | null — early bird days",
    "net_price": "number — net rate per room per night (THB)",
    "net_price_emerald": "number | null",
    "net_price_ruby": "number | null",
    "net_price_topaz": "number | null",
    "promo_book_till": "string | null — promotion booking deadline",
    "promo_code": "string | null",
    "promo_note": "string | null — e.g. \"Minimum 3 Nights stay\"",
    "room_allotment": "\"Free Sales\" | \"On Request\" | null",
    "all_inclusive": "bool | null",
    "baby_cot": "bool | null — true if complimentary baby cot available",
    "cancellation_policy": "HTML string with full cancellation policy",
    "cancellation_policy_net": "number | null — penalty amount",
    "early_check_in": "string | null",
    "child_policy": "HTML string with full child policy",
    "child_share_bed_abf": "number | null — child sharing bed with breakfast charge (THB)",
    "child_extra_bed_abf": "number | null — child extra bed + breakfast (THB)",
    "extra_bed_abf": "number | null — adult extra bed + breakfast (THB)",
    "extra_bed_no_abf": "number | null — adult extra bed without breakfast (THB)",
    "full_board": "number | null — full board supplement per adult per day (THB)",
    "half_board": "number | null — half board supplement per adult per day (THB)",
    "hotel_extra_fees": "string | null",
    "room_name": "string — room type name e.g. \"Deluxe Room\"",
    "hotel_transfer": "HTML string | null — airport transfer info and rates",
    "late_check_out": "string | null",
    "meals_and_info": "HTML string — comprehensive notes: supplements, offers, policies",
    "tags": "[] always empty array",
    "action": "\"insert\" always",
}


def _build_extract_prompt(hotel_id: str, room_id_map: dict) -> str:
    schema_lines = "\n".join(
        f'  "{k}": {v}' for k, v in _EXTRACT_COLS_SCHEMA.items()
    )

    room_id_section = ""
    if room_id_map:
        room_id_section = "\n\nROOM ID MAPPING (use these exact IDs):\n" + "\n".join(
            f'  "{name}" → room_id: "{rid}"' for name, rid in room_id_map.items()
        )
        room_id_section += f'\n  hotel_id for all rows: "{hotel_id}"'
    else:
        room_id_section = f'\n\nUse hotel_id: "{hotel_id}" for all rows. Set room_id to null (user will fill manually).'

    return f"""Extract ALL rate rows from this hotel contract PDF.

Return a JSON array where each object has EXACTLY these keys:
{{
{schema_lines}
}}
{room_id_section}

Important extraction notes:
- Create one row per (room_type × season_period) combination
- For "meals_and_info" field: compile ALL special offers, supplements, early bird, minimum stay, long stay, honeymoon, transfer, meal plans into one comprehensive HTML block
- For "cancellation_policy" field: write HTML covering the specific season's cancellation rules
- For "child_policy" field: write HTML with all child age bracket rules for this room type
- net_price = Room with Breakfast (RB) rate where available; fallback to Room Only (RO) if RB not available

Return ONLY the JSON array. Start your response with [ and end with ]"""


def generate_excel_from_pdf(
    pdf_bytes: bytes,
    api_key: str,
    hotel_id: str = "",
    room_id_map: dict | None = None,
) -> tuple[bytes | None, str]:
    """
    Extract contract data from PDF and return (xlsx_bytes, error_message).
    error_message is empty string on success.
    """
    client = _client(api_key)
    room_id_map = room_id_map or {}

    try:
        response = client.messages.create(
            model=MODEL,
            max_tokens=4096,
            system=_EXTRACT_SYSTEM,
            messages=[
                {
                    "role": "user",
                    "content": [
                        _pdf_block(pdf_bytes),
                        {"type": "text", "text": _build_extract_prompt(hotel_id, room_id_map)},
                    ],
                }
            ],
        )

        raw = response.content[0].text.strip()

        # Strip accidental markdown fences
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[1].rsplit("```", 1)[0].strip()

        rows = json.loads(raw)
        if not isinstance(rows, list):
            rows = [rows]

        xlsx_bytes = _rows_to_excel(rows, hotel_id)
        return xlsx_bytes, ""

    except json.JSONDecodeError as e:
        return None, f"AI returned invalid JSON: {e}"
    except anthropic.AuthenticationError:
        return None, "API Key ไม่ถูกต้อง"
    except anthropic.RateLimitError:
        return None, "โควต้า API เต็ม — กรุณาลองใหม่"
    except Exception as e:
        return None, str(e)


def _rows_to_excel(rows: list[dict], hotel_id: str) -> bytes:
    """Convert extracted row dicts into a formatted Excel workbook (bytes)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "AI"

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill_system = PatternFill("solid", fgColor="1F4E79")
    header_fill_extract = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill_system if col_name in SYSTEM_COLS else header_fill_extract

    # ── Data rows ─────────────────────────────────────────────────────────────
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_font = Font(name="Arial", size=10)
    wrap = Alignment(vertical="top", wrap_text=True)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = None
            if col_name == "_id":
                value = None  # system auto-fill
            elif col_name == "hotel_id":
                value = hotel_id or row_data.get("hotel_id")
            elif col_name == "room_id":
                value = row_data.get("room_id")
            elif col_name == "status":
                value = True
            elif col_name == "created_date":
                value = "=NOW()"
            elif col_name == "edited_date":
                value = None
            else:
                value = row_data.get(col_name)

            # Convert date strings to datetime objects
            if isinstance(value, str) and col_name in ("start_date", "end_date", "promo_book_till"):
                try:
                    value = datetime.strptime(value[:10], "%Y-%m-%d")
                except ValueError:
                    pass

            # Serialize lists/dicts to JSON string
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = border
            cell.alignment = wrap

            # Date formatting
            if col_name in ("start_date", "end_date") and isinstance(value, datetime):
                cell.number_format = "DD/MM/YYYY"

    # ── Column widths ─────────────────────────────────────────────────────────
    width_map = {
        "_id": 10, "hotel_id": 12, "room_id": 14, "status": 8,
        "created_date": 14, "edited_date": 14, "start_date": 12,
        "end_date": 12, "refundable": 10, "abf": 12, "contract_type": 16,
        "cutoff_date": 10, "hotel_supplier": 28, "important_message": 32,
        "min_nights_stay": 14, "min_advance_days": 16, "net_price": 12,
        "net_price_emerald": 16, "net_price_ruby": 14, "net_price_topaz": 14,
        "promo_book_till": 16, "promo_code": 14, "promo_note": 26,
        "room_allotment": 14, "all_inclusive": 12, "baby_cot": 10,
        "cancellation_policy": 48, "cancellation_policy_net": 22,
        "early_check_in": 14, "child_policy": 48, "child_share_bed_abf": 20,
        "child_extra_bed_abf": 20, "extra_bed_abf": 16, "extra_bed_no_abf": 18,
        "full_board": 12, "half_board": 12, "hotel_extra_fees": 26,
        "room_name": 22, "hotel_transfer": 48, "late_check_out": 14,
        "meals_and_info": 60, "tags": 10, "action": 10,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_map.get(col_name, 16)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "G2"

    # ── Save to bytes ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Feature 2b — Audit mode: compare existing Excel to PDF ────────────────────
def generate_excel_audit_report(
    pdf_bytes: bytes,
    excel_bytes: bytes,
    api_key: str,
    focus_list: list[str],
) -> str:
    """Non-streaming version that returns the full audit markdown."""
    return "".join(stream_recheck_analysis(pdf_bytes, excel_bytes, api_key, focus_list))
