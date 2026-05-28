"""
excel_generator.py — Generate a styled Excel workbook from the AI JSON data.
Layout matches the user's Google Sheets template (yellow header / blue period rows / red condition cells).
"""
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from utils_compare import clean_date_period_format



# ─── Style constants ──────────────────────────────────────────────────────────
_YELLOW = PatternFill("solid", fgColor="FFFF00")
_BLUE   = PatternFill("solid", fgColor="4A86E8")
_RED    = PatternFill("solid", fgColor="CC0000")
_LGRAY  = PatternFill("solid", fgColor="F3F4F6")

_WHITE_BOLD  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
_BLACK_BOLD  = Font(bold=True, name="Calibri", size=10)
_BLACK_NORM  = Font(name="Calibri", size=10)
_RED_BOLD_BIG = Font(color="CC0000", bold=True, name="Calibri", size=12)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _cell(ws, row, col, value="", fill=None, font=None, align=None, border=None):
    """Write a cell and optionally apply styles. Returns the cell."""
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill      = fill
    if font:   c.font      = font
    if align:  c.alignment = align
    if border: c.border    = border
    return c


def _border_row(ws, row, num_cols=8):
    """Apply thin border to every cell in a row."""
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).border = _THIN


def _clean_price(val):
    """Extract numeric value from strings like '1,500 THB' or '1,500.50'.
    Returns None for non-numeric values like 'N/A' or 'on request only'."""
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    # Remove commas and extract first valid float/int pattern
    import re
    s = str(val).replace(',', '').strip()
    if not s: return None
    
    # Detect explicit non-numeric markers
    s_lower = s.lower()
    if s_lower in ('n/a', '-', '--', 'na', 'tba', 'tbd') or 'on request' in s_lower or 'not available' in s_lower:
        return None
    
    # Try to find a number
    match = re.search(r"[-+]?\d*\.\d+|\d+", s)
    if match:
        try: 
            return float(match.group())
        except (ValueError, TypeError): 
            return None
    return None


def _fmt_nums(text: str) -> str:
    """Add comma separators to numbers >= 1,000 inside text strings.
    Skips numbers that follow date-related words (in, on, date, dates, year, etc.)
    because those are likely years, not prices.
    e.g. '1500 THB' -> '1,500 THB'
         'in 2025'  -> 'in 2025'  (unchanged)
         'dates in 2025-2026' -> unchanged
    """
    import re
    # Pattern to detect date-context keywords immediately before the number
    _DATE_PREFIX = re.compile(
        r'\b(?:date[s]?|in|on|year[s]?|by|from|until|since|period|'
        r'booking|staying|check-?in|check-?out|arrival|departure|valid|effective)\s*$',
        re.IGNORECASE
    )

    def _replace(m):
        # Check the text *before* this match for date-related words
        prefix = text[:m.start()]
        if _DATE_PREFIX.search(prefix):
            return m.group(0)          # looks like a year — leave as-is
        return f"{int(m.group(0)):,}"  # price — add comma

    return re.sub(r'\b\d{4,}\b', _replace, text)


def _generate_revise_comparison_excel_data(wb, data: dict):
    ws = wb.active
    ws.title = "Comparison Report"

    # Column widths (A–I)
    widths = [45, 15, 15, 15, 10, 12, 10, 12, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    row = 1
    hotel_name = data.get("hotel_name", "HOTEL NAME")
    year_1     = data.get("year_1", "24/25")
    year_2     = data.get("year_2", "25/26")
    year_3     = data.get("year_3", "25/26 REVISE")

    # Header Row
    _cell(ws, row, 1, str(hotel_name).upper(), fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 2, str(year_1).upper(),     fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 3, str(year_2).upper(),     fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 4, str(year_3).upper(),     fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 5, f"{year_1} : {year_2} %", fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 6, "Diff",                  fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 7, f"{year_2} : {year_3} %", fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 8, "Diff",                  fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 9, "",                      fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    row += 1

    seasons = data.get("seasons", [])
    for season in seasons:
        season_name = season.get("season_name") or ""
        p1 = clean_date_period_format(season.get("period_1") or "")
        p2 = clean_date_period_format(season.get("period_2") or "")
        p3 = clean_date_period_format(season.get("period_3") or "")

        # Period Header (Blue)
        _cell(ws, row, 1, "",                      fill=_BLUE, border=_THIN)
        _cell(ws, row, 2, p1,                      fill=_BLUE, font=_WHITE_BOLD, align=_CENTER, border=_THIN)
        _cell(ws, row, 3, p2,                      fill=_BLUE, font=_WHITE_BOLD, align=_CENTER, border=_THIN)
        _cell(ws, row, 4, p3,                      fill=_BLUE, font=_WHITE_BOLD, align=_CENTER, border=_THIN)
        _cell(ws, row, 5, "",                      fill=_BLUE, border=_THIN)
        _cell(ws, row, 6, "",                      fill=_BLUE, border=_THIN)
        _cell(ws, row, 7, "",                      fill=_BLUE, border=_THIN)
        _cell(ws, row, 8, "",                      fill=_BLUE, border=_THIN)
        _cell(ws, row, 9, str(season_name).upper() if season_name else "",
              fill=_YELLOW if season_name else _LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        row += 1

        cond_raw = season.get("conditions") or []
        if isinstance(cond_raw, list):
            cond_lines = cond_raw
        else:
            cond_lines = str(cond_raw).split('\n')
        
        rooms = season.get("rooms") or []
        for i, rm in enumerate(rooms):
            name = str(rm.get("room_name") or "").upper()
            p1_val = rm.get("price_1", 0)
            p2_val = rm.get("price_2", 0)
            p3_val = rm.get("price_3", 0)
            
            price_1 = _clean_price(p1_val)
            price_2 = _clean_price(p2_val)
            price_3 = _clean_price(p3_val)

            price_1_disp = price_1 if price_1 is not None else "N/A"
            price_2_disp = price_2 if price_2 is not None else "N/A"
            price_3_disp = price_3 if price_3 is not None else "N/A"

            # Diff 1-2
            if price_1 is not None and price_2 is not None:
                diff_amt_12 = price_2 - price_1
                diff_pct_12 = (price_2 - price_1) / price_2 if price_2 > 0 else 0.0
                diff_amt_12_disp = diff_amt_12
                diff_pct_12_disp = diff_pct_12
            else:
                diff_amt_12 = 0
                diff_pct_12 = 0
                diff_amt_12_disp = "-"
                diff_pct_12_disp = "-"

            # Diff 2-3
            if price_2 is not None and price_3 is not None:
                diff_amt_23 = price_3 - price_2
                diff_pct_23 = (price_3 - price_2) / price_3 if price_3 > 0 else 0.0
                diff_amt_23_disp = diff_amt_23
                diff_pct_23_disp = diff_pct_23
            else:
                diff_amt_23 = 0
                diff_pct_23 = 0
                diff_amt_23_disp = "-"
                diff_pct_23_disp = "-"

            _cell(ws, row, 1, name, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            p1_write = int(price_1) if price_1 is not None and price_1 == int(price_1) else price_1_disp
            p2_write = int(price_2) if price_2 is not None and price_2 == int(price_2) else price_2_disp
            p3_write = int(price_3) if price_3 is not None and price_3 == int(price_3) else price_3_disp
            
            _cell(ws, row, 2, p1_write, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            _cell(ws, row, 3, p2_write, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            _cell(ws, row, 4, p3_write, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            if price_1 is not None: ws.cell(row=row, column=2).number_format = '#,##0'
            if price_2 is not None: ws.cell(row=row, column=3).number_format = '#,##0'
            if price_3 is not None: ws.cell(row=row, column=4).number_format = '#,##0'

            # Format Diff 1-2
            pct_cell_12 = _cell(ws, row, 5, diff_pct_12_disp, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            if isinstance(diff_pct_12_disp, float):
                pct_cell_12.number_format = '0.00%'
                if diff_pct_12 > 0: pct_cell_12.font = Font(color="CC0000", bold=True, name="Calibri", size=10)
                elif diff_pct_12 < 0: pct_cell_12.font = Font(color="16A34A", bold=True, name="Calibri", size=10)

            amt_cell_12 = _cell(ws, row, 6, diff_amt_12_disp, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            if isinstance(diff_amt_12_disp, float):
                amt_cell_12.number_format = '#,##0;-#,##0;0'
                if diff_amt_12 > 0: amt_cell_12.font = Font(color="CC0000", bold=True, name="Calibri", size=10)
                elif diff_amt_12 < 0: amt_cell_12.font = Font(color="16A34A", bold=True, name="Calibri", size=10)

            # Format Diff 2-3
            pct_cell_23 = _cell(ws, row, 7, diff_pct_23_disp, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            if isinstance(diff_pct_23_disp, float):
                pct_cell_23.number_format = '0.00%'
                if diff_pct_23 > 0: pct_cell_23.font = Font(color="CC0000", bold=True, name="Calibri", size=10)
                elif diff_pct_23 < 0: pct_cell_23.font = Font(color="16A34A", bold=True, name="Calibri", size=10)

            amt_cell_23 = _cell(ws, row, 8, diff_amt_23_disp, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            if isinstance(diff_amt_23_disp, float):
                amt_cell_23.number_format = '#,##0;-#,##0;0'
                if diff_amt_23 > 0: amt_cell_23.font = Font(color="CC0000", bold=True, name="Calibri", size=10)
                elif diff_amt_23 < 0: amt_cell_23.font = Font(color="16A34A", bold=True, name="Calibri", size=10)

            cond_text = cond_lines[i] if i < len(cond_lines) else ""
            _cell(ws, row, 9, cond_text, font=Font(color="CC0000", name="Calibri", size=10), align=_LEFT, border=_THIN)
            row += 1

        for i in range(len(rooms), len(cond_lines)):
            if not cond_lines[i].strip(): continue
            for c in range(1, 9): _cell(ws, row, c, "", border=_THIN)
            _cell(ws, row, 9, cond_lines[i], font=Font(color="CC0000", name="Calibri", size=10), align=_LEFT, border=_THIN)
            row += 1

    row += 1

    def _extract_notes(lines):
        if lines is None: return "", []
        if not isinstance(lines, list):
            lines = [str(lines)]
        clean = []
        notes = []
        for line in lines:
            if not line: continue
            line_str = str(line)
            if "[RED]" in line_str:
                notes.append(line_str.replace("[RED]", "").strip())
            else:
                clean.append(line_str)
        return "\n".join(clean), notes

    def _build_stacked_section_revise(title, key):
        nonlocal row
        items = data.get(key, [])
        if not items: return
        
        _cell(ws, row, 1, str(title).upper(), fill=_YELLOW, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        for c in range(2, 10): _cell(ws, row, c, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1
        
        for item in items:
            c1_raw = item.get("contract_1", [])
            c2_raw = item.get("contract_2", [])
            c3_raw = item.get("contract_3", [])
            
            c1_text, c1_notes = _extract_notes(c1_raw)
            c2_text, c2_notes = _extract_notes(c2_raw)
            c3_text, c3_notes = _extract_notes(c3_raw)
            all_notes = c1_notes + [n for n in c2_notes if n not in c1_notes] + [n for n in c3_notes if n not in c1_notes and n not in c2_notes]
            
            c1_text = _fmt_nums(c1_text)
            c2_text = _fmt_nums(c2_text)
            c3_text = _fmt_nums(c3_text)
            
            # Contract 1
            _cell(ws, row, 1, f"CONTRACT {year_1}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
            for c in range(2, 10): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            row += 1
            
            _cell(ws, row, 1, c1_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            for c in range(2, 10): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            ws.row_dimensions[row].height = max(18, 15 * (c1_text.count("\n") + 1))
            row += 1
            
            # Contract 2
            _cell(ws, row, 1, f"CONTRACT {year_2}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
            for c in range(2, 10): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            row += 1
            
            _cell(ws, row, 1, c2_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            for c in range(2, 10): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            ws.row_dimensions[row].height = max(18, 15 * (c2_text.count("\n") + 1))
            row += 1

            # Contract 3
            _cell(ws, row, 1, f"CONTRACT {year_3}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
            for c in range(2, 10): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            row += 1
            
            _cell(ws, row, 1, c3_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            for c in range(2, 10): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            ws.row_dimensions[row].height = max(18, 15 * (c3_text.count("\n") + 1))
            row += 1
            
            # Notes
            if all_notes:
                notes_text = "\n".join(all_notes)
                _cell(ws, row, 1, notes_text, font=Font(color="CC0000", bold=True, name="Calibri", size=10), align=_LEFT, border=_THIN)
                for c in range(2, 10): _cell(ws, row, c, "", border=_THIN)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
                ws.row_dimensions[row].height = max(18, 15 * (notes_text.count("\n") + 1))
                row += 1
                
        row += 1

    def _build_sidebyside_section_revise(title, key):
        nonlocal row
        items = data.get(key, [])
        if not items: return
        
        _cell(ws, row, 1, str(title).upper(), fill=_YELLOW, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        for c in range(2, 10): _cell(ws, row, c, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1
        
        # Headers: 2 cols per contract + 3 cols for comparison
        _cell(ws, row, 1, f"CONTRACT {year_1}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        _cell(ws, row, 2, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        
        _cell(ws, row, 3, f"CONTRACT {year_2}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        _cell(ws, row, 4, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

        _cell(ws, row, 5, f"CONTRACT {year_3}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        _cell(ws, row, 6, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

        _cell(ws, row, 7, "COMPARISON / CHANGES", fill=PatternFill("solid", fgColor="D0E0F0"), font=Font(color="000080", bold=True, name="Calibri", size=10), align=_CENTER, border=_THIN)
        _cell(ws, row, 8, "", border=_THIN)
        _cell(ws, row, 9, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
        row += 1

        for item in items:
            c1_raw = item.get("contract_1", [])
            c2_raw = item.get("contract_2", [])
            c3_raw = item.get("contract_3", [])
            diff = str(item.get("diff_summary") or "SAME")
            diff2 = str(item.get("diff_summary_2") or "SAME")
            
            c1_text, c1_notes = _extract_notes(c1_raw)
            c2_text, c2_notes = _extract_notes(c2_raw)
            c3_text, c3_notes = _extract_notes(c3_raw)
            all_notes = c1_notes + [n for n in c2_notes if n not in c1_notes] + [n for n in c3_notes if n not in c1_notes and n not in c2_notes]
            
            c1_text = _fmt_nums(c1_text)
            c2_text = _fmt_nums(c2_text)
            c3_text = _fmt_nums(c3_text)

            # Contract 1
            _cell(ws, row, 1, c1_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            _cell(ws, row, 2, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

            # Contract 2
            _cell(ws, row, 3, c2_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            _cell(ws, row, 4, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

            # Contract 3
            _cell(ws, row, 5, c3_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            _cell(ws, row, 6, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

            # Comparison
            diff_text = f"C1 vs C2: {diff}\nC2 vs C3: {diff2}"
            _cell(ws, row, 7, diff_text, font=Font(color="000080", bold=True, name="Calibri", size=10), align=_CENTER, border=_THIN)
            _cell(ws, row, 8, "", border=_THIN)
            _cell(ws, row, 9, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
            
            ws.row_dimensions[row].height = max(35, 15 * (max(c1_text.count("\n"), c2_text.count("\n"), c3_text.count("\n"), diff_text.count("\n")) + 1))
            row += 1
            
            if all_notes:
                notes_text = "\n".join(all_notes)
                _cell(ws, row, 1, notes_text, font=Font(color="CC0000", bold=True, name="Calibri", size=10), align=_CENTER, border=_THIN)
                for c in range(2, 10): _cell(ws, row, c, "", border=_THIN)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
                ws.row_dimensions[row].height = max(18, 15 * (notes_text.count("\n") + 1))
                row += 1
                
        row += 1

    _build_stacked_section_revise("EXTRA BED / EXTRA PERSON", "extra_bed")
    _build_sidebyside_section_revise("EARLY BIRD OFFER", "early_bird")
    _build_sidebyside_section_revise("BONUS NIGHT OFFER", "bonus_night")
    _build_sidebyside_section_revise("WELLBEING SANCTUARY POOL SUITE LONG STAY BENEFITS", "wellbeing")
    _build_sidebyside_section_revise("CANCELLATION", "cancellation")
    _build_sidebyside_section_revise("OTHER PROMOTIONS / CONDITIONS", "other_promotions")


def generate_comparison_excel(data: dict) -> bytes:
    """
    สร้าง Excel workbook จาก JSON ที่ได้จาก Gemini Compare Contract
    Input schema:
        data = {
            "hotel_name": str,
            "year_1": str,           # เช่น "24/25"
            "year_2": str,           # เช่น "25/26"
            "seasons": [...],        # list of season objects
            "extra_bed": [...],      # policy sections
            "early_bird": [...],
            "bonus_night": [...],
            "wellbeing": [...],
            "cancellation": [...]
        }
    Output: bytes ของ .xlsx file พร้อม download
    """
    wb = openpyxl.Workbook()
    
    # If year_3 is present, use the Revise Contract Mode generator
    if "year_3" in data:
        _generate_revise_comparison_excel_data(wb, data)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    ws = wb.active
    ws.title = "Comparison Report"

    # Column widths (A–F)
    widths = [45, 25, 25, 10, 12, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    row = 1
    hotel_name = data.get("hotel_name", "HOTEL NAME")
    year_1     = data.get("year_1", "24/25")
    year_2     = data.get("year_2", "25/26")

    # Header Row
    hotel_name_upper = str(hotel_name).upper()
    year_1_upper     = str(year_1).upper()
    year_2_upper     = str(year_2).upper()
    _cell(ws, row, 1, hotel_name_upper, fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 2, year_1_upper,     fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 3, year_2_upper,     fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 4, "%",              fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 5, "",               fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 6, "",               fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    row += 1

    seasons = data.get("seasons", [])
    for season in seasons:
        season_name = season.get("season_name") or ""
        p1 = clean_date_period_format(season.get("period_1") or "")
        p2 = clean_date_period_format(season.get("period_2") or "")

        # Period Header (Blue)
        _cell(ws, row, 1, "",                      fill=_BLUE, border=_THIN)
        _cell(ws, row, 2, p1,                      fill=_BLUE, font=_WHITE_BOLD, align=_CENTER, border=_THIN)
        _cell(ws, row, 3, p2,                      fill=_BLUE, font=_WHITE_BOLD, align=_CENTER, border=_THIN)
        _cell(ws, row, 4, "",                      fill=_BLUE, border=_THIN)
        _cell(ws, row, 5, "",                      fill=_BLUE, border=_THIN)
        _cell(ws, row, 6, str(season_name).upper() if season_name else "",
              fill=_YELLOW if season_name else _LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        row += 1

        # Conditions for the season can go to Col F on the first room row
        cond_raw = season.get("conditions") or []
        if isinstance(cond_raw, list):
            cond_lines = cond_raw
        else:
            cond_lines = str(cond_raw).split('\n')
        
        rooms = season.get("rooms") or []
        for i, rm in enumerate(rooms):
            # Room name: uppercase
            name = str(rm.get("room_name") or "").upper()
            p1_val = rm.get("price_1", 0)
            p2_val = rm.get("price_2", 0)
            
            price_1 = _clean_price(p1_val)
            price_2 = _clean_price(p2_val)

            price_1_disp = price_1 if price_1 is not None else "N/A"
            price_2_disp = price_2 if price_2 is not None else "N/A"

            if price_1 is not None and price_2 is not None:
                diff_amt = price_2 - price_1
                # สูตรที่ถูกต้อง: (A2-A1)/A2 คือ ใช้ราคาปีใหม่ (price_2) เป็นตัวหาร
                if price_2 > 0:
                    diff_pct = (price_2 - price_1) / price_2
                elif price_2 == 0 and price_1 > 0:
                    diff_pct = -1.0  # -100%
                else:
                    diff_pct = 0.0
                diff_amt_disp = diff_amt
                diff_pct_disp = diff_pct
            else:
                diff_amt = 0
                diff_pct = 0
                diff_amt_disp = "-"
                diff_pct_disp = "-"

            _cell(ws, row, 1, name, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            # Write as int so #,##0 format shows commas correctly (1500 → 1,500)
            p1_write = int(price_1) if price_1 is not None and price_1 == int(price_1) else price_1_disp
            p2_write = int(price_2) if price_2 is not None and price_2 == int(price_2) else price_2_disp
            _cell(ws, row, 2, p1_write, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            _cell(ws, row, 3, p2_write, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            if price_1 is not None: ws.cell(row=row, column=2).number_format = '#,##0'
            if price_2 is not None: ws.cell(row=row, column=3).number_format = '#,##0'

            pct_cell = _cell(ws, row, 4, diff_pct_disp, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            if isinstance(diff_pct_disp, float):
                pct_cell.number_format = '0.00%'
                if diff_pct > 0: pct_cell.font = Font(color="CC0000", bold=True, name="Calibri", size=10)
                elif diff_pct < 0: pct_cell.font = Font(color="16A34A", bold=True, name="Calibri", size=10)

            amt_cell = _cell(ws, row, 5, diff_amt_disp, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            if isinstance(diff_amt_disp, float):
                amt_cell.number_format = '#,##0;-#,##0;0'
                if diff_amt > 0: amt_cell.font = Font(color="CC0000", bold=True, name="Calibri", size=10)
                elif diff_amt < 0: amt_cell.font = Font(color="16A34A", bold=True, name="Calibri", size=10)
            
            cond_text = cond_lines[i] if i < len(cond_lines) else ""
            _cell(ws, row, 6, cond_text, font=Font(color="CC0000", name="Calibri", size=10), align=_LEFT, border=_THIN)
            row += 1

        # If there are remaining conditions, print them in col F with empty room cells
        for i in range(len(rooms), len(cond_lines)):
            if not cond_lines[i].strip(): continue
            _cell(ws, row, 1, "", border=_THIN)
            _cell(ws, row, 2, "", border=_THIN)
            _cell(ws, row, 3, "", border=_THIN)
            _cell(ws, row, 4, "", border=_THIN)
            _cell(ws, row, 5, "", border=_THIN)
            _cell(ws, row, 6, cond_lines[i], font=Font(color="CC0000", name="Calibri", size=10), align=_LEFT, border=_THIN)
            row += 1

    row += 1

    def _extract_notes(lines):
        """Returns (clean_text, notes_list)"""
        if lines is None: return "", []
        if not isinstance(lines, list):
            lines = [str(lines)]
        clean = []
        notes = []
        for line in lines:
            if not line: continue
            line_str = str(line)
            if "[RED]" in line_str:
                notes.append(line_str.replace("[RED]", "").strip())
            else:
                clean.append(line_str)
        return "\n".join(clean), notes

    def _build_stacked_section(title, key):
        nonlocal row
        items = data.get(key, [])
        if not items: return
        
        _cell(ws, row, 1, str(title).upper(), fill=_YELLOW, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        
        for item in items:
            c1_raw = item.get("contract_1", [])
            c2_raw = item.get("contract_2", [])
            
            c1_text, c1_notes = _extract_notes(c1_raw)
            c2_text, c2_notes = _extract_notes(c2_raw)
            all_notes = c1_notes + [n for n in c2_notes if n not in c1_notes]
            # ใส่ลูกน้ำในตัวเลข >= 1,000 ที่อยู่ในข้อความ
            c1_text = _fmt_nums(c1_text)
            c2_text = _fmt_nums(c2_text)
            
            # Contract 1
            _cell(ws, row, 1, f"CONTRACT {year_1}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
            for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
            
            _cell(ws, row, 1, c1_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = max(18, 15 * (c1_text.count("\n") + 1))
            row += 1
            
            # Contract 2
            _cell(ws, row, 1, f"CONTRACT {year_2}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
            for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
            
            _cell(ws, row, 1, c2_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = max(18, 15 * (c2_text.count("\n") + 1))
            row += 1
            
            # Notes
            if all_notes:
                notes_text = "\n".join(all_notes)
                _cell(ws, row, 1, notes_text, font=Font(color="CC0000", bold=True, name="Calibri", size=10), align=_LEFT, border=_THIN)
                for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                ws.row_dimensions[row].height = max(18, 15 * (notes_text.count("\n") + 1))
                row += 1
                
        row += 1

    def _build_sidebyside_section(title, key):
        nonlocal row
        items = data.get(key, [])
        if not items: return
        
        _cell(ws, row, 1, str(title).upper(), fill=_YELLOW, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        
        # Headers
        _cell(ws, row, 1, f"CONTRACT {year_1}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        _cell(ws, row, 2, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        
        _cell(ws, row, 3, "COMPARISON / CHANGES", fill=PatternFill("solid", fgColor="D0E0F0"), font=Font(color="000080", bold=True, name="Calibri", size=10), align=_CENTER, border=_THIN)
        _cell(ws, row, 4, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

        _cell(ws, row, 5, f"CONTRACT {year_2}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        _cell(ws, row, 6, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        row += 1

        for item in items:
            c1_raw = item.get("contract_1", [])
            c2_raw = item.get("contract_2", [])
            diff = str(item.get("diff_summary") or "SAME")
            
            c1_text, c1_notes = _extract_notes(c1_raw)
            c2_text, c2_notes = _extract_notes(c2_raw)
            all_notes = c1_notes + [n for n in c2_notes if n not in c1_notes]
            # ใส่ลูกน้ำในตัวเลข >= 1,000 ที่อยู่ในข้อความ
            c1_text = _fmt_nums(c1_text)
            c2_text = _fmt_nums(c2_text)

            _cell(ws, row, 1, c1_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            _cell(ws, row, 2, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

            _cell(ws, row, 3, diff, font=Font(color="000080", bold=True, name="Calibri", size=10), align=_CENTER, border=_THIN)
            _cell(ws, row, 4, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

            _cell(ws, row, 5, c2_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            _cell(ws, row, 6, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            
            ws.row_dimensions[row].height = max(25, 15 * (max(c1_text.count("\n"), c2_text.count("\n")) + 1))
            row += 1
            
            # Notes span the full width below the side-by-side cells
            if all_notes:
                notes_text = "\n".join(all_notes)
                _cell(ws, row, 1, notes_text, font=Font(color="CC0000", bold=True, name="Calibri", size=10), align=_CENTER, border=_THIN)
                for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                ws.row_dimensions[row].height = max(18, 15 * (notes_text.count("\n") + 1))
                row += 1
                
        row += 1

    # Extra Bed is stacked in the image
    _build_stacked_section("EXTRA BED / EXTRA PERSON", "extra_bed")
    
    # Early Bird, Bonus Night, Wellbeing, Cancellation, Other Promotions are side-by-side
    _build_sidebyside_section("EARLY BIRD OFFER", "early_bird")
    _build_sidebyside_section("BONUS NIGHT OFFER", "bonus_night")
    _build_sidebyside_section("WELLBEING SANCTUARY POOL SUITE LONG STAY BENEFITS", "wellbeing")
    _build_sidebyside_section("CANCELLATION", "cancellation")
    _build_sidebyside_section("OTHER PROMOTIONS / CONDITIONS", "other_promotions")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
